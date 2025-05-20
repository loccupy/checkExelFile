from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

from libs.all_data_object import PathToFile


def change_cell_color(cell_address):
    path = PathToFile.path
    try:
        # Открываем существующий файл
        workbook = load_workbook(path)
        sheet = workbook.active

        # Создаем заливку нужного цвета
        fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

        # Применяем заливку к ячейке
        sheet[cell_address].fill = fill

        # Сохраняем изменения
        workbook.save(path)
        print(f"Цвет ячейки {cell_address} в файле {path} изменен на красный")
    except FileNotFoundError:
        print("Файл не найден")
        exit()
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
        exit()


def change_cell_color_orange(cell_address):
    path = PathToFile.path
    try:
        # Открываем существующий файл
        workbook = load_workbook(path)
        sheet = workbook.active

        # Создаем заливку нужного цвета
        fill = PatternFill(start_color='FFA000', end_color='FFA000', fill_type='solid')

        # Применяем заливку к ячейке
        sheet[cell_address].fill = fill

        # Сохраняем изменения
        workbook.save(path)
        print(f"Цвет ячейки {cell_address} в файле {path} изменен на оранжевый")
    except FileNotFoundError:
        print("Файл не найден")
        exit()
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")
        exit()